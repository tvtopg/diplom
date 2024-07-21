from docxtpl import DocxTemplate #для работы с  документами Word
import openpyxl #для работы с файлами Excel
from mergeExcelBooks import merge
# from db import databaseQuery

import sys #для работы с операционной системой.
import os

# excelFileName = 'bdV.xlsx' # имя файла Excel, из которого будут браться данные.  +
# sheetName = 'lines' # имя листа в файле Excel, с которого будут браться данные.
# fileNameinField = 'Fname' # имя поля, в котором указывается имя файла для сохранения
# wordFileName = 'TemplateRPV.docx' #имя шаблона документа Word
# saveFolder = '' #папка для сохранения

# wb = openpyxl.load_workbook(excelFileName, data_only=True) #Открывается файл
# sheet=wb[sheetName] #Выбирается лист
# context={}

# excelFileName = 'excelBooks/Шишкина А. А.xlsx' # имя файла Excel, из которого будут браться данные.  +
# sheetName = 'Философия' # имя листа в файле Excel, с которого будут браться данные.
# fileNameinField = 'Fname' # имя поля, в котором указывается имя файла для сохранения
# wordFileName = 'excelBooks/TemplateRPVNew.docx' #имя шаблона документа Word
# saveFolder = 'Work programs' #папка для сохранения

# wb = openpyxl.load_workbook(excelFileName, data_only=True) #Открывается файл
# sheet=wb[sheetName] #Выбирается лист
# context={}

excelFileName = 'excelBooks/Шишкина А. А.xlsx' # имя файла Excel, из которого будут браться данные.  +
sheetName = 'Философия' # имя листа в файле Excel, с которого будут браться данные.
fileNameinField = 'Fname' # имя поля, в котором указывается имя файла для сохранения
wordFileName = 'excelBooks/TemplateRPVNew.docx' #имя шаблона документа Word
saveFolder = 'Work programs' #папка для сохранения

wb = openpyxl.load_workbook(excelFileName, data_only=True) #Открывается файл
sheet=wb[sheetName] #Выбирается лист
context={}

def parseOneNesting(sheet, column,row): #принимает в parsecells ключ, строку и колонку
    i = 0
    data = []
    while (sheet.cell(row=row+i, column=1).value==None or i==0) and (sheet.cell(row=row+i, column=column).value!=None):
        data.append(sheet.cell(row=row+i, column=column).value)
        i += 1    
    return data

def parseTwoNesting(sheet, column,row): #принимает в parsecells ключ, строку и колонку
    i = 0
    data = []
    while (sheet.cell(row=row, column=column+i).value!=None):
        datadata = {'value':sheet.cell(row=row, column=column+i).value}
        j = 1 # перескочить на нижний уровень списка
        dataitems = []
        while (sheet.cell(row=row+j, column=1).value==None or j==0) and (sheet.cell(row=row+j, column=column+i).value!=None):
            str = splitString(sheet.cell(row=row+j, column=column+i).value)
            dataitems.extend(str)
            # print(dataitems, end='\n')
            j += 1
        datadata['elements'] = dataitems    
        data.append(datadata)
        i += 1
    # elem[fieldname[0:-2]] = data
    # print(data)
    return data

def parseTable(sheet, column,row): #принимает в parsecells ключ, строку и колонку
    # print(sheet.cell(row=row, column=column).value)
    i = 0
    data = {}
    while sheet.cell(row=row, column=column+i).value!=None:
        dataitems = []
        j = 2

        while (sheet.cell(row=row+j, column=1).value==None) and (sheet.cell(row=row+j, column=column+i).value!=None):
            dataitems.append(sheet.cell(row=row+j, column=column+i).value)
            j+=1
        data[sheet.cell(row=row, column=column+i).value] = dataitems
        i+=1
        # print(data)
    # print(data, end='\n')
    # print(end='\n')
    return data

def splitString(inputString):
    return inputString.split('\n')

# fieldsnames=[] # пустой список
# nc = 1 # номер колонки

def excelContext(excelFileName, sheetName): # В прошлом был filling
    excelFileName = excelFileName # имя файла Excel, из которого будут браться данные.  +
    sheetName = sheetName # имя листа в файле Excel, с которого будут браться данные.


    wb = openpyxl.load_workbook(excelFileName, data_only=True) #Открывается файл
    sheet=wb[sheetName] #Выбирается лист
    context={}
    for row in sheet: # цикл по строчкам в листе Excel
        if row[0].value==None:
            continue
        key = str(row[0].value)
        if key[-2:]=='**':
            context[key[:-2]] = parseTwoNesting(sheet, column=3, row=row[0].row)
            # print(key[:-2], context[key[:-2]])
            continue
        if key[-1:]=='*':
            context[key[:-1]] = parseOneNesting(sheet, column=3, row=row[0].row)
            # print(key[:-1],context[key[:-1]])
            continue
        if key=='Table':
            context[key] = context.update(parseTable(sheet, column=3, row=row[0].row))
            continue
        context[row[0].value] = row[2].value
    
    # doc = DocxTemplate(wordFileName)
    # doc.render(context)
    # doc.save(str(saveFolder + '/' + context[fileNameinField])+".docx")
    return context







