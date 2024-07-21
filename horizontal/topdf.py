from docxtpl import DocxTemplate #для работы с  документами Word
import openpyxl #для работы с файлами Excel
excelFileName = 'bd.xlsx' # имя файла Excel
sheetName = 'lines' # имя листа в файле Excel
fileNameinField = 'Fname' # имя поля, в котором указывается имя файла для сохранения
wordFileName = 'TemplateRP.docx' #имя шаблона документа Word
import sys
import os
#import comtypes.client #для работы с приложением Microsoft Word
wdFormatPDF = 17

wb = openpyxl.load_workbook(excelFileName, data_only=True)
sheet=wb[sheetName]

def parsecells(fieldname,column,row): #принимает в parsecells ключ, строку и колонку
    elem = {}
    if fieldname==None:
        return elem
    if fieldname[-1]!='*': # последняя буква ключа #(выполняется при отсутсвии звёздочек)
        elem[fieldname] = sheet.cell(row=row, column=column).value
    elif fieldname[-2:]!='**': # последние две буквы ключа (выполняется при *)
        i = 0
        data = []
        while sheet.cell(row=row+i, column=column).value!=None and (i==0 or sheet.cell(row=row+i, column=1).value==None):
            data.append(sheet.cell(row=row+i, column=column).value)
            i += 1
        elem[fieldname[0:-1]] = data #вычитается последняя символ (*) и поулчается ключ
    else:  #(выполняется при **)
        i = 0
        data = []
        while (sheet.cell(row=row+i, column=column).value!=None or sheet.cell(row=row+i+ii, column=column+1).value!=None) and (i==0 or sheet.cell(row=row+i, column=1).value==None):
            datadata = {'value':sheet.cell(row=row+i, column=column).value}
            ii = 0
            dataitems = []
            while sheet.cell(row=row+i+ii, column=column+1).value!=None and (ii==0 or sheet.cell(row=row+i+ii, column=column).value==None):
                dataitems.append(sheet.cell(row=row+i+ii, column=column+1).value)
                ii += 1
            datadata['elements'] = dataitems
            data.append(datadata)
            print(data)
            i += ii
        elem[fieldname[0:-2]] = data
    return elem
    
fieldsnames=[] # Сюда записываются ключи (N, Tshort, Tfull, Tprof, Tkaf)
rn = 1 # номер предмета
for row in sheet: # цикл по строкам в листе Excel
    # print(str(row[0].value),str(rn-1), end='\n')

    # пройтись и собрать все ключи
    if rn==1 and row[0].value == 'N':# Если rn равно 1 и значение в первой ячейке равно 'N', то добавляем значения ячеек в fieldsnames.
        for cellObj in row: # цикл по ячекам в текущей строке
            fieldsnames.append(cellObj.value) # Значения ячеек добавляются в список 
    else:

        # Флажок, типо предмет 1, 2 ... (пропуск пустых строчек)
        if str(row[0].value)!=str(rn-1): # 1!=0, 2!=1
            continue
        # Проходимся по данным 
        context={}
        for cellObj in row: # Цикл по строке
            # print(fieldsnames[cellObj.column-1],cellObj.row, cellObj.column, end='\n')
            if cellObj.value==None:
                continue
            # передаёт в parsecells ключ, строку и колонку
            context.update(parsecells(fieldsnames[cellObj.column-1],row=cellObj.row,column=cellObj.column))
        #print(context)
        doc = DocxTemplate(wordFileName)
        # doc.render(context)
        # doc.save(str(context[fileNameinField])+".docx")

        #не надо
        #in_file = os.path.abspath (str(context[fileNameinField])+".docx")
        #out_file = os.path.abspath(str(context[fileNameinField])+".pdf")
        #word = comtypes.client.CreateObject('Word.Application')
        #doc = word.Documents.Open(in_file)
        #doc.SaveAs(out_file, FileFormat=wdFormatPDF)
        #doc.Close()
        #word.Quit()
    rn += 1


# Этот код представляет собой функцию `parsecells`,
# которая используется для анализа ячеек в таблице
# (предположительно, в формате Excel или подобном).
#  Давайте разберем каждую строчку кода:

# 1. `def parsecells(fieldname, column, row):` - Определение функции `parsecells`
# с тремя параметрами: `fieldname` (ключ), `column` (колонка) и `row` (строка).

# 2. `elem = {}` - Создание пустого словаря `elem`, который будет заполняться
#  данными из ячеек.

# 3. `if fieldname == None:` - Проверка, равен ли `fieldname` `None`. Если да,
#  то функция возвращает пустой словарь `elem`.

# 4. `if fieldname[-1] != '*':` - Проверка, не заканчивается ли `fieldname`
#  символом '*'. Если нет, то код выполняется для обычной ячейки.

#     - `elem[fieldname] = sheet.cell(row=row, column=column).value`
#     - Присваивание значению ячейки в указанной строке `row` и колонке `column`
#       ключа `fieldname` в словаре `elem`.

# 5. `elif fieldname[-2:] != '**':` - Если `fieldname` заканчивается одной '*'. 
# Это означает, что нужно обработать последовательность ячеек в столбце.

#     - `i = 0` - Инициализация переменной `i` для отслеживания текущей строки.
#     - `data = []` - Инициализация пустого списка `data` для хранения данных из последовательности ячеек.

#     - `while sheet.cell(row=row+i, column=column).value != None and (i==0 or sheet.cell(row=row+i, column=1).value==None):` - Цикл, который продолжается, пока значение в текущей ячейке не является `None` и либо это первая итерация (i==0), либо значение в первом столбце текущей строки равно `None`.

#         - `data.append(sheet.cell(row=row+i, column=column).value)` - Добавление значения текущей ячейки в список `data`.
#         - `i += 1` - Увеличение `i` для перехода к следующей строке.

#     - `elem[fieldname[0:-1]] = data` - Присваивание значению ключа `fieldname` в словаре `elem` списка `data`.

# 6. `else:` - Если `fieldname` заканчивается двумя '*'. Это означает, что нужно обработать группу ячеек и их подгруппы.

#     - `i = 0` - Инициализация переменной `i` для отслеживания текущей строки.
#     - `data = []` - Инициализация пустого списка `data` для хранения данных из группы ячеек.

#     - `while (sheet.cell(row=row+i, column=column).value != None or sheet.cell(row=row+i+ii, column=column+1).value != None) and (i==0 or sheet.cell(row=row+i, column=1).value==None):` - Цикл, который продолжается, пока значение в текущей ячейке не является `None` или значение в следующей ячейке второго столбца не является `None`, и либо это первая итерация (i==0), либо значение в первом столбце текущей строки равно `None`.

#         - `datadata = {'value':sheet.cell(row=row+i, column=column).value}` - Создание словаря `datadata` с ключом 'value', содержащим значение текущей ячейки.
#         - `ii = 0` - Инициализация переменной `ii` для отслеживания текущей строки второго столбца.
#         - `dataitems = []` - Инициализация пустого списка `dataitems` для хранения данных из подгруппы ячеек.

#         - `while sheet.cell(row=row+i+ii, column=column+1).value != None and (ii==0 or sheet.cell(row=row+i+ii, column=column).value==None):` - Цикл, который продолжается, пока значение в текущей ячейке второго столбца не является `None` и либо это первая итерация (ii==0), либо значение в первом столбце текущей строки второго столбца равно `None`.

#             - `dataitems.append(sheet.cell(row=row+i+ii, column=column+1).value)` - Добавление значения текущей ячейки второго столбца в список `dataitems`.
#             - `ii += 1` - Увеличение `ii` для перехода к следующей строке второго столбца.

#         - `datadata['elements'] = dataitems` - Присваивание значению ключа 'elements' в словаре `datadata` списка `dataitems`.
#         - `data.append(datadata)` - Добавление словаря `datadata` в список `data`.
#         - `i += ii` - Увеличение `i` на `ii` для перехода к следующей строке после обработки подгруппы ячеек.

#     - `elem[fieldname[0:-2]] = data` - Присваивание значению ключа `fieldname` в словаре `elem` списка `data`.

# 7. `return elem` - Возврат словаря `elem` из функции.

# Эта функция служит для обработки различных типов данных в ячейках таблицы и возвращает словарь с полученными данными.